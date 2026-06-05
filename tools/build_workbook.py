"""
build_workbook.py — Regenerate the master budget workbook from plan + ledger.

Produces data/Yash_Budget.xlsx with four sheets:
  1. Plan            — every head, per-occurrence amount, frequency, monthly-equiv
  2. Dashboard       — head-wise Budget vs Actual vs Variance for ONE month
  3. Monthly Actuals — matrix: heads (rows) x months (cols) of actual spend
  4. Transactions    — the raw ledger

Also writes a CSV mirror of the Monthly Actuals matrix for easy cloud upload.

Usage:
    python tools/build_workbook.py                # dashboard = latest month in ledger
    python tools/build_workbook.py 2026-06        # dashboard for a specific month
"""
import csv
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from lib_budget import (load_plan, load_ledger, load_income, load_balances,
                        monthly_equivalent, actuals_by_head_month, dashboard_matrix,
                        fy_months, current_month, income_matrix, pnl_matrix,
                        balances_matrix, ROOT)

OUT_XLSX = ROOT / "data" / "Yash_Budget.xlsx"
OUT_CSV = ROOT / "data" / "monthly_actuals.csv"

# ---- styling -------------------------------------------------------------
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
OVER_FILL = PatternFill("solid", fgColor="F8CBAD")   # red-ish: overspent
UNDER_FILL = PatternFill("solid", fgColor="C6EFCE")  # green-ish: under budget
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center")
RIGHT = Alignment(horizontal="right")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '#,##0'


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def months_in_ledger(ledger):
    return sorted({t["month"] for t in ledger["transactions"]})


def build_plan_sheet(wb, plan):
    ws = wb.create_sheet("Plan")
    headers = ["Particulars", "Per Occurrence", "Frequency", "Due Months",
               "Monthly Equivalent", "Remarks"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    total_me = 0.0
    for h in plan["heads"]:
        me = monthly_equivalent(plan, h)
        total_me += me
        due = ", ".join(str(m) for m in h.get("due_months", [])) or "-"
        ws.append([h["particulars"], h["budget"], h["frequency"], due,
                   round(me, 2), h["remarks"]])
    ws.append(["TOTAL (bare-minimum monthly)", "", "", "", round(total_me, 2), ""])
    last = ws.max_row
    for c in range(1, len(headers) + 1):
        ws.cell(row=last, column=c).fill = TOTAL_FILL
        ws.cell(row=last, column=c).font = BOLD
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=2).number_format = MONEY
        ws.cell(row=r, column=5).number_format = MONEY
    _autosize(ws, [34, 16, 14, 14, 18, 38])
    ws.freeze_panes = "A2"
    return total_me


def build_dashboard(wb, plan, ledger, months):
    """Multi-month FY dashboard. `months` is the FY axis e.g.
    ['2026-04','2026-05','2026-06']."""
    ws = wb.create_sheet("Dashboard", 0)  # first sheet
    headers, body, total = dashboard_matrix(plan, ledger, months)
    ncols = len(headers)
    n = len(months)
    monthvar_col = ncols        # <Mon> vs Budget — last column
    status_col = ncols - 1      # Status
    ytdvar_col = ncols - 2      # YTD Variance
    curmonth_col = 2 + n        # the latest month's actual column (1-based)
    span = f"{months[0][:7]} to {months[-1][:7]}" if months else "no data yet"

    ws.append([f"BUDGET DASHBOARD  —  FY 2026-27  ({span})"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F4E78")
    ws.append([])
    ws.append(headers)
    hrow = ws.max_row
    _style_header(ws, hrow, ncols)

    for row in body:
        ws.append(row)
        r = ws.max_row
        # YTD accrual view: flag the Status (+ YTD Variance) cell
        status = row[status_col - 1]
        if status == "OVER":
            ws.cell(row=r, column=status_col).fill = OVER_FILL
            ws.cell(row=r, column=ytdvar_col).fill = OVER_FILL
        elif status in ("Under", "On Track"):
            ws.cell(row=r, column=status_col).fill = UNDER_FILL
        # Current-month view: red if this month's spend is over the monthly
        # budget, green if under — but only when there was spend this month.
        cur_actual = row[curmonth_col - 1]
        if cur_actual:
            ws.cell(row=r, column=monthvar_col).fill = (
                OVER_FILL if row[monthvar_col - 1] > 0 else UNDER_FILL)

    ws.append(total)
    last = ws.max_row
    for c in range(1, ncols + 1):
        ws.cell(row=last, column=c).fill = TOTAL_FILL
        ws.cell(row=last, column=c).font = BOLD

    for r in range(hrow + 1, ws.max_row + 1):
        for c in range(2, ncols + 1):  # every numeric col except Status
            if c == status_col:
                continue
            ws.cell(row=r, column=c).number_format = MONEY
            ws.cell(row=r, column=c).alignment = RIGHT
    # widths: Particulars, Monthly Budget, per-month, YTD, Budget, YTD Var, Status, Mon-var
    widths = [34, 15] + [12] * len(months) + [13, 16, 13, 16, 17]
    _autosize(ws, widths)
    ws.freeze_panes = "C4"


def build_monthly_actuals(wb, plan, ledger, months):
    ws = wb.create_sheet("Monthly Actuals")
    abm = actuals_by_head_month(ledger)
    headers = ["Particulars", "Monthly Budget"] + months
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    csv_rows = [headers]
    col_totals = defaultdict(float)
    for h in plan["heads"]:
        me = monthly_equivalent(plan, h)
        row = [h["particulars"], round(me, 2)]
        for m in months:
            v = abm.get(h["id"], {}).get(m, 0.0)
            row.append(round(v, 2))
            col_totals[m] += v
        ws.append(row)
        csv_rows.append(row)
    total_row = ["TOTAL", round(sum(monthly_equivalent(plan, h) for h in plan["heads"]), 2)]
    total_row += [round(col_totals[m], 2) for m in months]
    ws.append(total_row)
    csv_rows.append(total_row)

    last = ws.max_row
    for c in range(1, len(headers) + 1):
        ws.cell(row=last, column=c).fill = TOTAL_FILL
        ws.cell(row=last, column=c).font = BOLD
    for r in range(2, ws.max_row + 1):
        for c in range(2, len(headers) + 1):
            ws.cell(row=r, column=c).number_format = MONEY
    _autosize(ws, [34, 16] + [13] * len(months))
    ws.freeze_panes = "C2"

    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(csv_rows)


def build_balances(wb, balances):
    ws = wb.create_sheet("Balances", 2)  # after Dashboard, P&L
    headers, body, net = balances_matrix(balances)
    ws.append(["ACCOUNT BALANCES  —  latest known position"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F4E78")
    ws.append([])
    ws.append(headers)
    _style_header(ws, ws.max_row, len(headers))
    for r in body:
        ws.append(r)
        cell = ws.cell(row=ws.max_row, column=3)
        cell.fill = UNDER_FILL if r[2] >= 0 else OVER_FILL
    ws.append(net)
    last = ws.max_row
    for c in range(1, len(headers) + 1):
        ws.cell(row=last, column=c).fill = TOTAL_FILL
        ws.cell(row=last, column=c).font = BOLD
    for r in range(4, ws.max_row + 1):
        ws.cell(row=r, column=3).number_format = MONEY
        ws.cell(row=r, column=3).alignment = RIGHT
    ws.append([])
    ws.append(["Liabilities (credit-card outstanding) shown negative. Investments to be added."])
    _autosize(ws, [34, 12, 16, 12, 28])
    return ws


def build_income(wb, income, months):
    ws = wb.create_sheet("Income")
    headers, body, total = income_matrix(income, months)
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    for r in body:
        ws.append(r)
    ws.append(total)
    last = ws.max_row
    for c in range(1, len(headers) + 1):
        ws.cell(row=last, column=c).fill = TOTAL_FILL
        ws.cell(row=last, column=c).font = BOLD
    # below-the-line inflows that are NOT counted as income
    below = ([("OTHER INFLOWS - not income (asset sale / capital refund)", income.get("other_inflows", []))]
             + [("PENDING CLASSIFICATION (not counted yet)", income.get("pending_classification", []))])
    for title, items in below:
        if not items:
            continue
        ws.append([])
        ws.append([title])
        ws.cell(row=ws.max_row, column=1).font = BOLD
        for r in items:
            ws.append([r["date"], r["source"], r.get("category", ""),
                       round(float(r["amount"])), "", r.get("note", "")])
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=4).number_format = MONEY
    _autosize(ws, [12, 30, 20, 14, 22, 40])
    ws.freeze_panes = "A2"


def build_pnl(wb, ledger, income, months):
    ws = wb.create_sheet("P&L", 1)  # right after Dashboard
    headers, body = pnl_matrix(ledger, income, months)
    ws.append([f"PROFIT & LOSS  —  FY 2026-27  (income vs recorded expenses)"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F4E78")
    ws.append([])
    ws.append(headers)
    _style_header(ws, ws.max_row, len(headers))
    for r in body:
        ws.append(r)
        cell = ws.cell(row=ws.max_row, column=1)
        if "Net" in r[0]:
            cell.font = BOLD
    for r in range(4, ws.max_row + 1):
        for c in range(2, len(headers) + 1):
            ws.cell(row=r, column=c).number_format = MONEY
            ws.cell(row=r, column=c).alignment = RIGHT
    _autosize(ws, [26] + [13] * (len(headers) - 1))
    ws.append([])
    ws.append(["Note: expenses are statements recorded so far (partial as accounts are added)."])
    return ws


def build_transactions(wb, ledger):
    ws = wb.create_sheet("Transactions")
    headers = ["Date", "Month", "Description", "Amount", "Account", "Head", "Note"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    for t in sorted(ledger["transactions"], key=lambda x: x["date"]):
        ws.append([t["date"], t["month"], t["description"], t["amount"],
                   t["account"], t["head_id"], t["note"]])
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=4).number_format = MONEY
    _autosize(ws, [12, 10, 36, 14, 14, 18, 24])
    ws.freeze_panes = "A2"


def main(as_of=None):
    """as_of is the 'YYYY-MM' the FY axis runs up to (default: current month).
    The dashboard always shows April -> as_of for FY 2026-27."""
    plan = load_plan()
    ledger = load_ledger()
    income = load_income()
    balances = load_balances()
    if as_of is None:
        as_of = current_month()
    months = fy_months(as_of)

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet
    build_dashboard(wb, plan, ledger, months)
    build_pnl(wb, ledger, income, months)
    build_balances(wb, balances)
    build_plan_sheet(wb, plan)
    build_monthly_actuals(wb, plan, ledger, months)
    build_income(wb, income, months)
    build_transactions(wb, ledger)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_CSV}")
    print(f"Dashboard FY axis: {', '.join(months)} | "
          f"transactions: {len(ledger['transactions'])}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else None)
